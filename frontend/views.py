from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.db.models import Count, Q
from django.utils import timezone
from datetime import date, timedelta
import calendar
from .models import Course, Group, Student, MarketingSurvey, StudentLog, LessonTime, Branch, Room, Role, Position, Employee, Attendance, AbsenceReason
from .forms import LoginForm, CourseForm, GroupForm, StudentCreateForm, StudentEditForm, MarketingSurveyForm, FreezeForm, RemoveFromGroupForm, AddToGroupForm, LessonTimeForm, BranchForm, RoomForm, PositionForm, EmployeeForm


def login_view(request):
    if request.user.is_authenticated:
        try:
            if request.user.employee_profile.role and request.user.employee_profile.role.name == "O'qituvchi":
                return redirect("teacher_dashboard")
        except:
            pass
        return redirect("dashboard")
    form = LoginForm()
    if request.method == "POST":
        form = LoginForm(request.POST)
        if form.is_valid():
            user = authenticate(
                request,
                username=form.cleaned_data["phone"],
                password=form.cleaned_data["password"],
            )
            if user:
                login(request, user)
                try:
                    if user.employee_profile.role and user.employee_profile.role.name == "O'qituvchi":
                        return redirect("teacher_dashboard")
                except:
                    pass
                return redirect("dashboard")
            messages.error(request, "Telefon raqam yoki parol noto'g'ri")
    return render(request, "login.html", {"form": form})


def logout_view(request):
    logout(request)
    return redirect("login")


@login_required(login_url="login")
def teacher_dashboard(request):
    try:
        employee = request.user.employee_profile
        if not employee.role or employee.role.name != "O'qituvchi":
            messages.error(request, "Siz o'qituvchi emassiz!")
            return redirect("login")
    except:
        messages.error(request, "Siz o'qituvchi emassiz!")
        return redirect("login")

    from datetime import datetime, time, date
    now = datetime.now()
    current_time = now.time()
    current_weekday = now.weekday()
    weekday_map = {
        0: "dushanba", 1: "seshanba", 2: "chorshanba",
        3: "payshanba", 4: "juma", 5: "shanba", 6: "yakshanba"
    }
    today_uz = weekday_map[current_weekday]

    today_date = date.today()
    weekdays_uz = {
        0: "Dushanba", 1: "Seshanba", 2: "Chorshanba",
        3: "Payshanba", 4: "Juma", 5: "Shanba", 6: "Yakshanba"
    }
    months_uz = {
        1: "Yanvar", 2: "Fevral", 3: "Mart", 4: "Aprel",
        5: "May", 6: "Iyun", 7: "Iyul", 8: "Avgust",
        9: "Sentabr", 10: "Oktabr", 11: "Noyabr", 12: "Dekabr"
    }
    today_display = f"{weekdays_uz[today_date.weekday()]}, {today_date.day} {months_uz[today_date.month]} {today_date.year}"

    day_filter = request.GET.get("day", "")
    date_filter = request.GET.get("date", "")

    all_teacher_groups = Group.objects.filter(
        teacher=employee, status__in=["aktiv", "kutilyotgan"]
    ).select_related("course", "room").prefetch_related("lesson_times", "students").annotate(
        student_count=Count("students")
    )

    selected_date_display = ""
    if date_filter:
        try:
            from datetime import datetime as dt
            parsed = dt.strptime(date_filter, "%Y-%m-%d")
            wd = weekday_map[parsed.weekday()]
            groups = all_teacher_groups.filter(lesson_times__days__contains=wd)
            day_filter = wd
            selected_date_display = f"{weekdays_uz[parsed.weekday()]}, {parsed.day} {months_uz[parsed.month]} {parsed.year}"
        except:
            groups = all_teacher_groups
    elif day_filter:
        groups = all_teacher_groups.filter(lesson_times__days__contains=day_filter)
    else:
        groups = all_teacher_groups

    groups = groups.distinct().order_by("name")

    today_count = 0
    active_count = 0
    total_students = 0

    group_list = []
    for g in groups:
        if g.is_date_overdue():
            group_list.append({
                "group": g,
                "student_count": g.students.count(),
                "lesson_display": "",
                "status": "expired",
                "nearest_time": None,
            })
            continue

        lesson_times = list(g.lesson_times.all())
        status = "kutilmoqda"
        lesson_display = ""
        nearest_time = None

        students_count = g.students.count()
        total_students += students_count

        for lt in lesson_times:
            days_list = [d.strip().lower() for d in lt.days.split(",") if d.strip()]
            day_matches = today_uz in days_list
            if day_filter:
                day_matches = day_filter in days_list

            day_matches_for_today = today_uz in days_list

            if nearest_time is None or (lt.start_time and (nearest_time is None or lt.start_time < nearest_time)):
                nearest_time = lt.start_time

            if day_matches_for_today:
                today_count += 1
                if lt.start_time <= current_time <= lt.end_time:
                    status = "active"
                    active_count += 1
                elif lt.end_time < current_time:
                    if status != "active":
                        status = "finished"
                elif lt.start_time > current_time:
                    if status not in ("active", "finished"):
                        status = "upcoming"

            if day_filter and day_matches:
                lesson_display = f"{lt.get_days_display()} {lt.start_time.strftime('%H:%M')}-{lt.end_time.strftime('%H:%M')}"
            elif not day_filter:
                lesson_display = f"{lt.get_days_display()} {lt.start_time.strftime('%H:%M')}-{lt.end_time.strftime('%H:%M')}"

        group_list.append({
            "group": g,
            "student_count": students_count,
            "lesson_display": lesson_display or (str(lesson_times[0]) if lesson_times else ""),
            "status": status,
            "nearest_time": nearest_time,
        })

    active_groups = [g for g in group_list if g["status"] == "active"]
    upcoming_groups = [g for g in group_list if g["status"] == "upcoming"]
    finished_groups = [g for g in group_list if g["status"] == "finished"]
    pending_groups = [g for g in group_list if g["status"] == "kutilmoqda"]
    expired_groups = [g for g in group_list if g["status"] == "expired"]

    sorted_groups = active_groups + upcoming_groups + finished_groups + pending_groups + expired_groups

    total_groups = all_teacher_groups.count()

    return render(request, "teacher/dashboard.html", {
        "groups": sorted_groups,
        "employee": employee,
        "selected_day": day_filter,
        "selected_date": date_filter,
        "selected_date_display": selected_date_display,
        "total_groups": total_groups,
        "today_count": today_count,
        "active_count": active_count,
        "total_students": total_students,
        "today_display": today_display,
    })


@login_required(login_url="login")
def teacher_my_groups(request):
    try:
        employee = request.user.employee_profile
        if not employee.role or employee.role.name != "O'qituvchi":
            messages.error(request, "Siz o'qituvchi emassiz!")
            return redirect("login")
    except:
        messages.error(request, "Siz o'qituvchi emassiz!")
        return redirect("login")

    groups = Group.objects.filter(
        teacher=employee, status__in=["aktiv", "kutilyotgan"]
    ).select_related("course", "room").prefetch_related(
        "lesson_times", "students"
    ).annotate(student_count=Count("students")).distinct().order_by("name")

    from datetime import datetime, date
    now = datetime.now()
    current_time = now.time()
    current_weekday = now.weekday()
    weekday_map = {0:"dushanba",1:"seshanba",2:"chorshanba",3:"payshanba",4:"juma",5:"shanba",6:"yakshanba"}
    today_uz = weekday_map[current_weekday]

    group_list = []
    for g in groups:
        if g.is_date_overdue():
            group_list.append({"group":g,"student_count":g.students.count(),"lesson_display":"","status":"expired","nearest_time":None})
            continue
        lesson_times = list(g.lesson_times.all())
        status = "kutilmoqda"
        lesson_display = ""
        nearest_time = None
        for lt in lesson_times:
            days_list = [d.strip().lower() for d in lt.days.split(",") if d.strip()]
            if nearest_time is None or (lt.start_time and lt.start_time < nearest_time):
                nearest_time = lt.start_time
            if today_uz in days_list:
                if lt.start_time <= current_time <= lt.end_time:
                    status = "active"
                elif lt.end_time < current_time:
                    if status != "active":
                        status = "finished"
                elif lt.start_time > current_time:
                    if status not in ("active","finished"):
                        status = "upcoming"
            lesson_display = f"{lt.get_days_display()} {lt.start_time.strftime('%H:%M')}-{lt.end_time.strftime('%H:%M')}"
        group_list.append({"group":g,"student_count":g.students.count(),"lesson_display":lesson_display,"status":status,"nearest_time":nearest_time})

    return render(request, "teacher/my_groups.html", {
        "groups": group_list,
        "employee": employee,
    })


@login_required(login_url="login")
def admin_mobile_groups(request):
    is_admin = request.user.is_staff
    if not is_admin:
        emp = getattr(request.user, 'employee_profile', None)
        if emp is None or not emp.role or emp.role.name != "O'qituvchi":
            is_admin = True
    if not is_admin:
        messages.error(request, "Siz admin emassiz!")
        return redirect("login")

    groups = Group.objects.filter(
        status__in=["aktiv", "kutilyotgan"]
    ).select_related("course", "room", "teacher").prefetch_related(
        "lesson_times", "students"
    ).annotate(student_count=Count("students")).distinct().order_by("name")

    from datetime import datetime, date
    now = datetime.now()
    current_time = now.time()
    current_weekday = now.weekday()
    weekday_map = {0:"dushanba",1:"seshanba",2:"chorshanba",3:"payshanba",4:"juma",5:"shanba",6:"yakshanba"}
    today_uz = weekday_map[current_weekday]

    # Only today's groups
    groups = groups.filter(lesson_times__days__contains=today_uz)

    group_list = []
    for g in groups:
        if g.is_date_overdue():
            continue
        lesson_times = list(g.lesson_times.all())
        status = "kutilmoqda"
        lesson_display = ""
        nearest_time = None
        for lt in lesson_times:
            days_list = [d.strip().lower() for d in lt.days.split(",") if d.strip()]
            if nearest_time is None or (lt.start_time and lt.start_time < nearest_time):
                nearest_time = lt.start_time
            if today_uz in days_list:
                if lt.start_time <= current_time <= lt.end_time:
                    status = "active"
                elif lt.end_time < current_time:
                    if status != "active":
                        status = "finished"
                elif lt.start_time > current_time:
                    if status not in ("active","finished"):
                        status = "upcoming"
            lesson_display = f"{lt.get_days_display()} {lt.start_time.strftime('%H:%M')}-{lt.end_time.strftime('%H:%M')}"
        group_list.append({"group":g,"student_count":g.students.count(),"lesson_display":lesson_display,"status":status,"nearest_time":nearest_time})

    return render(request, "teacher/admin_mobile_groups.html", {
        "groups": group_list,
    })


@login_required(login_url="login")
def teacher_group_detail(request, pk):
    try:
        employee = request.user.employee_profile
    except:
        messages.error(request, "Siz o'qituvchi emassiz!")
        return redirect("login")

    group = get_object_or_404(
        Group.objects.select_related("course", "room").prefetch_related("lesson_times", "students"),
        pk=pk, teacher=employee
    )
    students = group.students.all().order_by("first_name")

    from datetime import date
    today_attendances = Attendance.objects.filter(group=group, date=date.today())
    attendance_map = {a.student_id: a.status for a in today_attendances}
    attendance_notes = {a.student_id: a.notes for a in today_attendances if a.notes}

    # find today's lesson time
    weekday_map = {0:"dushanba",1:"seshanba",2:"chorshanba",3:"payshanba",4:"juma",5:"shanba",6:"yakshanba"}
    today_uz = weekday_map[date.today().weekday()]
    today_lesson = group.lesson_times.filter(days__contains=today_uz).first()

    absence_reasons = AbsenceReason.objects.filter(is_active=True).order_by("order", "name")

    return render(request, "teacher/group_detail.html", {
        "group": group,
        "students": students,
        "employee": employee,
        "attendance_map": attendance_map,
        "attendance_notes": attendance_notes,
        "absence_reasons": absence_reasons,
        "today_lesson": today_lesson,
    })


from django.views.decorators.csrf import csrf_exempt
@login_required(login_url="login")
@csrf_exempt
def take_attendance(request):
    if request.method != "POST":
        return JsonResponse({"error": "POST required"}, status=405)
    import json
    from datetime import datetime, time as dt_time
    try:
        data = json.loads(request.body)
    except:
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    group_id = data.get("group_id")
    records = data.get("records", [])  # [{student_id, status}]

    try:
        employee = request.user.employee_profile
    except:
        employee = None

    is_admin = request.user.is_staff
    # Non-staff users with non-teacher role are also admins
    if not is_admin and (employee is None or not employee.role or employee.role.name != "O'qituvchi"):
        is_admin = True
    group = get_object_or_404(Group, pk=group_id)
    today = date.today()
    allow_dated = data.get("allow_dated", False)

    # Permission check
    if is_admin:
        # Admin can save attendance for any group
        pass
    else:
        # Teacher can only save for their own groups
        if employee is None or group.teacher_id != employee.id:
            return JsonResponse({"error": "Siz o'qituvchi emassiz!"}, status=403)
        # Teacher time restriction
        weekday_map = {0:"dushanba",1:"seshanba",2:"chorshanba",3:"payshanba",4:"juma",5:"shanba",6:"yakshanba"}
        today_uz = weekday_map[today.weekday()]
        now = datetime.now().time()
        allowed = False
        for lt in group.lesson_times.all():
            for d_name in lt.days.split(","):
                if d_name.strip() == today_uz:
                    if lt.start_time and lt.end_time:
                        if lt.start_time <= now <= lt.end_time:
                            allowed = True
                            break
            if allowed:
                break
        if not allowed:
            return JsonResponse({"error": "Dars vaqti ichida bo'lmaganda davomatni o'zgartira olmaysiz!"}, status=403)

    student_ids_in_records = set()
    created_by = ''
    if employee:
        created_by = f"{employee.first_name} {employee.last_name or ''}".strip()
    elif request.user.is_staff:
        emp = getattr(request.user, 'employee_profile', None)
        if emp:
            created_by = f"{emp.first_name} {emp.last_name or ''}".strip()
        if not created_by or created_by.replace('+', '').replace(' ', '').isdigit():
            created_by = request.user.get_full_name()
        if not created_by or created_by.replace('+', '').replace(' ', '').isdigit():
            created_by = "Admin"
    for rec in records:
        student_id = rec.get("student_id")
        status = rec.get("status", "present")
        rec_date = rec.get("date")
        if allow_dated and rec_date:
            try:
                rec_date = date.fromisoformat(rec_date)
            except:
                rec_date = today
        else:
            rec_date = today

        # Teacher can only save for today
        if not is_admin and rec_date != today:
            return JsonResponse({"error": "Faqat bugungi davomatni o'zgartira olasiz!"}, status=403)

        if status == "none":
            Attendance.objects.filter(
                group=group, student_id=student_id, date=rec_date
            ).delete()
            continue
        student_ids_in_records.add(student_id)
        notes = rec.get("notes", "") or ""
        Attendance.objects.update_or_create(
            group=group,
            student_id=student_id,
            date=rec_date,
            defaults={"status": status, "teacher": employee, "notes": notes, "created_by": created_by}
        )

    # Tegilmagan o'quvchilarni "Keldi" qilib saqlash
    if not allow_dated:
        for student in group.students.all():
            if student.id not in student_ids_in_records:
                Attendance.objects.update_or_create(
                    group=group,
                    student_id=student.id,
                    date=today,
                    defaults={"status": "present", "teacher": employee, "notes": "", "created_by": created_by}
                )

    return JsonResponse({"ok": True})



@login_required(login_url="login")
def teacher_attendance_desktop(request, pk):
    is_admin = request.user.is_staff
    if not is_admin:
        employee = getattr(request.user, 'employee_profile', None)
        if employee is None or not employee.role or employee.role.name != "O'qituvchi":
            is_admin = True

    if is_admin:
        group = get_object_or_404(
            Group.objects.select_related("course", "room", "teacher"),
            pk=pk
        )
        employee = getattr(request.user, 'employee_profile', None)
    else:
        try:
            employee = request.user.employee_profile
        except:
            messages.error(request, "Siz o'qituvchi emassiz!")
            return redirect("login")

        group = get_object_or_404(
            Group.objects.select_related("course", "room", "teacher"),
            pk=pk, teacher=employee
        )
    students = group.students.all().order_by("first_name")

    # Month/year from query string, default to current
    today = date.today()
    sel_year = int(request.GET.get("year", today.year))
    sel_month = int(request.GET.get("month", today.month))

    # Generate lesson dates for the selected month
    weekday_map_rev = {0:"dushanba",1:"seshanba",2:"chorshanba",3:"payshanba",4:"juma",5:"shanba",6:"yakshanba"}
    lesson_day_numbers = set()
    for lt in group.lesson_times.all():
        for d_name in lt.days.split(","):
            d_name = d_name.strip()
            for num, uz_name in weekday_map_rev.items():
                if uz_name == d_name:
                    lesson_day_numbers.add(num)

    _, last_day = calendar.monthrange(sel_year, sel_month)
    month_start = date(sel_year, sel_month, 1)
    month_end = date(sel_year, sel_month, last_day)

    group_start = group.start_date if group.start_date else month_start
    lesson_start = max(month_start, group_start)

    lesson_dates = []
    d = lesson_start
    while d <= month_end:
        if d.weekday() in lesson_day_numbers:
            lesson_dates.append(d)
        d += timedelta(days=1)

    # Build attendance matrix: {student_id: {date_str: status}}
    attendances = Attendance.objects.filter(group=group, date__gte=month_start, date__lte=month_end)
    att_matrix = {}
    att_notes = {}
    for a in attendances:
        sid = a.student_id
        if sid not in att_matrix:
            att_matrix[sid] = {}
        att_matrix[sid][a.date.isoformat()] = a.status
        if a.notes:
            if sid not in att_notes:
                att_notes[sid] = {}
            att_notes[sid][a.date.isoformat()] = a.notes

    # Stats for the month
    total = students.count()
    present_count = sum(1 for a in attendances if a.status == "present")
    absent_count = sum(1 for a in attendances if a.status == "absent")
    excused_count = sum(1 for a in attendances if a.status == "excused")
    boshqoldi_count = sum(1 for a in attendances if a.status == "boshqoldi")
    total_marked = present_count + absent_count + excused_count + boshqoldi_count

    def pct(n): return round((n / total * 100)) if total else 0

    # Build month options for selector
    months_uz = ["", "Yanvar", "Fevral", "Mart", "Aprel", "May", "Iyun", "Iyul", "Avgust", "Sentabr", "Oktabr", "Noyabr", "Dekabr"]

    # Check if teacher can edit (only during lesson time)
    from datetime import datetime
    can_edit = True if is_admin else False
    if not is_admin:
        weekday_map = {0:"dushanba",1:"seshanba",2:"chorshanba",3:"payshanba",4:"juma",5:"shanba",6:"yakshanba"}
        today_uz = weekday_map[today.weekday()]
        now = datetime.now().time()
        for lt in group.lesson_times.all():
            for d_name in lt.days.split(","):
                if d_name.strip() == today_uz:
                    if lt.start_time and lt.end_time:
                        if lt.start_time <= now <= lt.end_time:
                            can_edit = True
                            break
            if can_edit:
                break

    absence_reasons = AbsenceReason.objects.filter(is_active=True).order_by("order", "name")

    # Today's groups for mobile group picker (admin only)
    today_groups = []
    if is_admin:
        from datetime import datetime
        now = datetime.now()
        weekday_map = {0:"dushanba",1:"seshanba",2:"chorshanba",3:"payshanba",4:"juma",5:"shanba",6:"yakshanba"}
        today_uz = weekday_map[now.weekday()]
        qs = Group.objects.filter(status__in=["aktiv","kutilyotgan"], lesson_times__days__contains=today_uz).select_related("course","teacher").annotate(sc=Count("students")).distinct().order_by("name")
        for g in qs:
            today_groups.append({"id":g.id,"name":g.name,"course":g.course.name,"teacher_name":str(g.teacher),"sc":g.sc,"active":g.id==group.id})

    # Per-student latest absence reason (for the "Sabab" column)
    student_last_reason = {}
    for a in attendances.filter(status__in=['absent', 'excused']).exclude(notes__exact='').order_by('-date'):
        if a.student_id not in student_last_reason:
            student_last_reason[a.student_id] = a.notes

    # Per-student full attendance history
    status_labels = {'absent': 'Kelmadi', 'excused': 'Sababli', 'present': 'Keldi', 'boshqoldi': 'Davom olish'}
    student_att_history = {}
    for a in attendances.order_by('-created_at'):
        sid = a.student_id
        if sid not in student_att_history:
            student_att_history[sid] = []
        teacher_name = ''
        if a.teacher:
            teacher_name = (a.teacher.first_name or '') + ' ' + (a.teacher.last_name or '')
            teacher_name = teacher_name.strip()
        elif a.created_by:
            teacher_name = a.created_by
        from datetime import timezone as dt_timezone, timedelta as dt_timedelta
        tashkent_tz = dt_timezone(dt_timedelta(hours=5))
        day_names = {0:'Du',1:'Se',2:'Cho',3:'Pay',4:'Ju',5:'Sha',6:'Yak'}
        student_att_history[sid].append({
            'date': a.date.isoformat(),
            'datetime': a.created_at.astimezone(tashkent_tz).strftime('%d.%m.%Y %H:%M'),
            'day_name': day_names[a.date.weekday()],
            'status': status_labels.get(a.status, a.status),
            'notes': a.notes or '',
            'teacher': teacher_name,
        })

    return render(request, "teacher/attendance_desktop.html", {
        "group": group,
        "students": students,
        "employee": employee,
        "lesson_dates": lesson_dates,
        "att_matrix": att_matrix,
        "att_notes": att_notes,
        "student_last_reason": student_last_reason,
        "sel_year": sel_year,
        "sel_month": sel_month,
        "months_uz": months_uz,
        "years": range(2024, 2031),
        "today": today,
        "total": total,
        "present_count": present_count,
        "absent_count": absent_count,
        "excused_count": excused_count,
        "boshqoldi_count": boshqoldi_count,
        "total_marked": total_marked,
        "student_last_reason": student_last_reason,
        "student_att_history": student_att_history,
        "is_admin": is_admin,
        "can_edit": can_edit,
        "lessons_count": len(lesson_dates),
        "absence_reasons": absence_reasons,
        "today_groups": today_groups,
    })


@login_required(login_url="login")
def dashboard(request):
    try:
        if request.user.employee_profile.role and request.user.employee_profile.role.name == "O'qituvchi":
            return redirect("teacher_dashboard")
    except:
        pass
    course_count = Course.objects.count()
    group_count = Group.objects.count()
    student_count = Student.objects.count()
    survey_count = MarketingSurvey.objects.count()
    active_groups = Group.objects.filter(status="aktiv").count()
    pending_groups = Group.objects.filter(status="kutilyotgan").count()
    pending_student_count = Student.objects.filter(groups__isnull=True, status="kutilyotgan").count()
    recent_students = Student.objects.prefetch_related("groups").order_by("-created_at")[:5]
    survey_stats = MarketingSurvey.objects.annotate(
        student_count=Count("students")
    ).filter(student_count__gt=0).order_by("-student_count")
    return render(request, "dashboard.html", {
        "course_count": course_count,
        "group_count": group_count,
        "student_count": student_count,
        "survey_count": survey_count,
        "active_groups": active_groups,
        "pending_groups": pending_groups,
        "pending_student_count": pending_student_count,
        "recent_students": recent_students,
        "survey_stats": survey_stats,
    })


@login_required(login_url="login")
def course_list(request):
    courses = Course.objects.annotate(group_count=Count("groups")).order_by("-created_at")
    return render(request, "course/list.html", {"courses": courses})


@login_required(login_url="login")
def course_create(request):
    form = CourseForm()
    if request.method == "POST":
        form = CourseForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Kurs muvaffaqiyatli qo'shildi")
            return redirect("course_list")
    return render(request, "course/form.html", {"form": form, "title": "Kurs qo'shish"})


@login_required(login_url="login")
def course_update(request, pk):
    course = get_object_or_404(Course, pk=pk)
    form = CourseForm(instance=course)
    if request.method == "POST":
        form = CourseForm(request.POST, instance=course)
        if form.is_valid():
            form.save()
            messages.success(request, "Kurs muvaffaqiyatli yangilandi")
            return redirect("course_list")
    return render(request, "course/form.html", {"form": form, "title": "Kursni tahrirlash"})


@login_required(login_url="login")
def course_delete(request, pk):
    course = get_object_or_404(Course, pk=pk)
    if request.method == "POST":
        course.delete()
        messages.success(request, "Kurs muvaffaqiyatli o'chirildi")
        return redirect("course_list")
    return render(request, "course/delete.html", {"object": course, "title": "Kursni o'chirish"})


@login_required(login_url="login")
def group_list(request):
    status = request.GET.get("status")
    day = request.GET.get("day")
    day_type = request.GET.get("day_type")
    time_from = request.GET.get("time_from")
    time_to = request.GET.get("time_to")
    course_id = request.GET.get("course")
    teacher_id = request.GET.get("teacher")
    room_id = request.GET.get("room")
    search = request.GET.get("search")

    if status == "arxivlangan":
        status_filter = ["arxivlangan"]
        page_title = "Arxivlangan guruhlar"
    else:
        status_filter = ["aktiv", "kutilyotgan"]
        page_title = "Oddiy guruhlar"

    groups = Group.objects.select_related("course", "room", "teacher").prefetch_related(
        "lesson_times"
    ).annotate(
        student_count=Count("students")
    ).filter(status__in=status_filter)

    if day_type:
        groups = groups.filter(day_type=day_type)

    if day:
        groups = groups.filter(lesson_times__days__contains=day)

    if time_from:
        groups = groups.filter(lesson_times__start_time__gte=time_from)

    if time_to:
        groups = groups.filter(lesson_times__end_time__lte=time_to)

    if course_id:
        groups = groups.filter(course_id=course_id)

    if teacher_id:
        groups = groups.filter(teacher_id=teacher_id)

    if room_id:
        groups = groups.filter(room_id=room_id)

    if search:
        groups = groups.filter(name__icontains=search)

    groups = groups.distinct().order_by("-created_at")

    courses = Course.objects.all()
    teachers = Employee.objects.filter(role__name="O'qituvchi")
    rooms = Room.objects.all()

    return render(request, "group/list.html", {
        "groups": groups,
        "page_title": page_title,
        "courses": courses,
        "teachers": teachers,
        "rooms": rooms,
    })


@login_required(login_url="login")
def group_create(request):
    form = GroupForm()
    if request.method == "POST":
        form = GroupForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Guruh muvaffaqiyatli qo'shildi")
            return redirect("group_list")
    return render(request, "group/form.html", {"form": form, "title": "Guruh qo'shish"})


@login_required(login_url="login")
def group_update(request, pk):
    group = get_object_or_404(Group, pk=pk)
    form = GroupForm(instance=group)
    if request.method == "POST":
        form = GroupForm(request.POST, instance=group)
        if form.is_valid():
            form.save()
            messages.success(request, "Guruh muvaffaqiyatli yangilandi")
            return redirect("group_list")
    return render(request, "group/form.html", {"form": form, "title": "Guruhni tahrirlash"})


@login_required(login_url="login")
def group_delete(request, pk):
    group = get_object_or_404(Group, pk=pk)
    if request.method == "POST":
        group.delete()
        messages.success(request, "Guruh muvaffaqiyatli o'chirildi")
        return redirect("group_list")
    return render(request, "group/delete.html", {"object": group, "title": "Guruhni o'chirish"})


@login_required(login_url="login")
def group_extend(request, pk):
    group = get_object_or_404(Group, pk=pk)
    if request.method == "POST":
        days = request.POST.get("days")
        if days and days.isdigit() and int(days) > 0:
            from datetime import timedelta
            group.end_date += timedelta(days=int(days))
            group.save()
            messages.success(request, f"Guruh muddati {days} kunga uzaytirildi")
            return redirect("group_detail", pk=group.pk)
        messages.error(request, "Kunlar sonini to'g'ri kiriting")
    return render(request, "group/extend.html", {"group": group})


@login_required(login_url="login")
def student_list(request):
    from django.db.models import Count, Q

    search = request.GET.get("search", "").strip()
    status_filter = request.GET.get("status", "")
    group_id = request.GET.get("group", "")

    students = Student.objects.prefetch_related("groups", "marketing_survey").annotate(
        group_count=Count("groups")
    ).order_by("-created_at")

    if search:
        q_filter = Q(first_name__icontains=search) | Q(last_name__icontains=search)
        digits_only = ''.join(c for c in search if c.isdigit())
        if digits_only:
            q_filter |= Q(phone__icontains=digits_only)
        students = students.filter(q_filter)

    if status_filter == "aktiv":
        students = students.filter(groups__isnull=False).exclude(status="chiqarilgan").exclude(frozen_until__gte=date.today())
    elif status_filter == "muzlatilgan":
        students = students.filter(frozen_until__gte=date.today())
    elif status_filter == "kutilyotgan":
        students = students.filter(groups__isnull=True, status="kutilyotgan")
    elif status_filter == "chiqarilgan":
        students = students.filter(status="chiqarilgan")

    if group_id:
        students = students.filter(groups__id=group_id)

    students = students.distinct()

    groups = Group.objects.filter(status__in=["aktiv", "kutilyotgan"]).order_by("name")

    return render(request, "student/list.html", {
        "students": students,
        "groups": groups,
        "active_filters": {
            "search": search,
            "status": status_filter,
            "group": group_id,
        },
    })


@login_required(login_url="login")
def student_export_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse

    students = Student.objects.prefetch_related("groups", "marketing_survey").annotate(
        group_count=Count("groups")
    ).order_by("-created_at")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "O'quvchilar"

    headers = [
        "#", "Ism", "Familya", "Telefon", "Guruhlar",
        "Guruhlar soni", "Holat", "Tug'ilgan sana",
        "Ota ism", "Ota nomer", "Ona ism", "Ona nomer",
        "Qo'shilgan sana"
    ]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    for i, s in enumerate(students, 1):
        groups_str = ", ".join(g.name for g in s.groups.all()) if s.groups.exists() else "-"
        if s.is_frozen:
            status = "Muzlatilgan"
        elif s.groups.exists():
            status = "Aktiv"
        elif s.status == "chiqarilgan":
            status = "Chiqarilgan"
        else:
            status = "Kutilyotgan"
        row = [
            i, s.first_name, s.last_name, s.phone, groups_str,
            s.group_count, status,
            s.birth_date.strftime("%d.%m.%Y") if s.birth_date else "-",
            s.father_full_name or "-", s.father_phone or "-",
            s.mother_full_name or "-", s.mother_phone or "-",
            s.created_at.strftime("%d.%m.%Y") if s.created_at else "-",
        ]
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=i+1, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 35
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 14
    ws.column_dimensions['I'].width = 22
    ws.column_dimensions['J'].width = 20
    ws.column_dimensions['K'].width = 22
    ws.column_dimensions['L'].width = 20
    ws.column_dimensions['M'].width = 14

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="oquvchilar.xlsx"'
    wb.save(response)
    return response


@login_required(login_url="login")
def student_export_csv(request):
    import csv
    from django.http import HttpResponse

    students = Student.objects.prefetch_related("groups", "marketing_survey").annotate(
        group_count=Count("groups")
    ).order_by("-created_at")

    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="oquvchilar.csv"'
    response.write('\ufeff')

    writer = csv.writer(response)
    writer.writerow([
        "#", "Ism", "Familya", "Telefon", "Guruhlar",
        "Guruhlar soni", "Holat", "Tug'ilgan sana",
        "Ota ism", "Ota nomer", "Ona ism", "Ona nomer",
        "Qo'shilgan sana"
    ])

    for i, s in enumerate(students, 1):
        groups_str = ", ".join(g.name for g in s.groups.all()) if s.groups.exists() else "-"
        if s.is_frozen:
            status = "Muzlatilgan"
        elif s.groups.exists():
            status = "Aktiv"
        elif s.status == "chiqarilgan":
            status = "Chiqarilgan"
        else:
            status = "Kutilyotgan"
        writer.writerow([
            i, s.first_name, s.last_name, s.phone, groups_str,
            s.group_count, status,
            s.birth_date.strftime("%d.%m.%Y") if s.birth_date else "-",
            s.father_full_name or "-", s.father_phone or "-",
            s.mother_full_name or "-", s.mother_phone or "-",
            s.created_at.strftime("%d.%m.%Y") if s.created_at else "-",
        ])
    return response


def _write_csv(response, headers, rows):
    import csv
    response.write('\ufeff')
    writer = csv.writer(response)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    return response


@login_required(login_url="login")
def group_export_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    groups = Group.objects.select_related("course", "teacher", "room").prefetch_related("lesson_times").annotate(student_count=Count("students"))
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Guruhlar"
    headers = ["#","Nomi","Kurs","O'qituvchi","Xona","Kunlar","Vaqt","Talabalar","Holat","Boshlanish","Tugash"]
    hf = Font(bold=True,color="FFFFFF",size=11); hfl = PatternFill(start_color="2563EB",end_color="2563EB",fill_type="solid")
    tb = Border(left=Side(style='thin',color='D1D5DB'),right=Side(style='thin',color='D1D5DB'),top=Side(style='thin',color='D1D5DB'),bottom=Side(style='thin',color='D1D5DB'))
    for c,h in enumerate(headers,1):
        cell = ws.cell(row=1,column=c,value=h); cell.font = hf; cell.fill = hfl; cell.alignment = Alignment(horizontal='center',vertical='center'); cell.border = tb
    for i,g in enumerate(groups,1):
        lt = g.lesson_times.first(); days = lt.days if lt else ""; time = f"{lt.start_time}—{lt.end_time}" if lt else ""
        teacher = f"{g.teacher.first_name} {g.teacher.last_name}" if g.teacher else "-"
        row = [i,g.name,g.course.name if g.course else "-",teacher,g.room.name if g.room else "-",days,time,g.student_count,g.get_status_display(),g.start_date.strftime('%d.%m.%Y') if g.start_date else '-',g.end_date.strftime('%d.%m.%Y') if g.end_date else '-']
        for c,v in enumerate(row,1): cell = ws.cell(row=i+1,column=c,value=v); cell.border = tb; cell.alignment = Alignment(vertical='center')
    for col,w in [(1,5),(2,25),(3,18),(4,22),(5,14),(6,14),(7,14),(8,10),(9,14),(10,14),(11,14)]: ws.column_dimensions[chr(64+col)].width = w
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="guruhlar.xlsx"'; wb.save(response); return response

@login_required(login_url="login")
def group_export_csv(request):
    groups = Group.objects.select_related("course","teacher","room").prefetch_related("lesson_times").annotate(student_count=Count("students"))
    headers = ["#","Nomi","Kurs","O'qituvchi","Xona","Kunlar","Vaqt","Talabalar","Holat","Boshlanish","Tugash"]
    rows = []
    for i,g in enumerate(groups,1):
        lt = g.lesson_times.first(); days = lt.days if lt else ""; time = f"{lt.start_time}—{lt.end_time}" if lt else ""
        teacher = f"{g.teacher.first_name} {g.teacher.last_name}" if g.teacher else "-"
        rows.append([i,g.name,g.course.name if g.course else "-",teacher,g.room.name if g.room else "-",days,time,g.student_count,g.get_status_display(),g.start_date.strftime('%d.%m.%Y') if g.start_date else '-',g.end_date.strftime('%d.%m.%Y') if g.end_date else '-'])
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="guruhlar.csv"'
    return _write_csv(response, headers, rows)

@login_required(login_url="login")
def employee_export_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    employees = Employee.objects.prefetch_related("branches").select_related("position","role").all().order_by("-created_at")
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Xodimlar"
    headers = ["#","Ism","Familya","Telefon","Lavozim","Rol","Filiallar","Qo'shilgan sana"]
    hf = Font(bold=True,color="FFFFFF",size=11); hfl = PatternFill(start_color="2563EB",end_color="2563EB",fill_type="solid")
    tb = Border(left=Side(style='thin',color='D1D5DB'),right=Side(style='thin',color='D1D5DB'),top=Side(style='thin',color='D1D5DB'),bottom=Side(style='thin',color='D1D5DB'))
    for c,h in enumerate(headers,1): cell = ws.cell(row=1,column=c,value=h); cell.font = hf; cell.fill = hfl; cell.alignment = Alignment(horizontal='center',vertical='center'); cell.border = tb
    for i,e in enumerate(employees,1):
        branches = ", ".join(b.name for b in e.branches.all()) if e.branches.exists() else "-"
        row = [i,e.first_name,e.last_name,e.phone,e.position.name if e.position else "-",e.role.name if e.role else "-",branches,e.created_at.strftime('%d.%m.%Y') if e.created_at else '-']
        for c,v in enumerate(row,1): cell = ws.cell(row=i+1,column=c,value=v); cell.border = tb; cell.alignment = Alignment(vertical='center')
    for col,w in [(1,5),(2,18),(3,18),(4,20),(5,22),(6,18),(7,22),(8,14)]: ws.column_dimensions[chr(64+col)].width = w
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="xodimlar.xlsx"'; wb.save(response); return response

@login_required(login_url="login")
def employee_export_csv(request):
    employees = Employee.objects.prefetch_related("branches").select_related("position","role").all().order_by("-created_at")
    headers = ["#","Ism","Familya","Telefon","Lavozim","Rol","Filiallar","Qo'shilgan sana"]
    rows = []
    for i,e in enumerate(employees,1):
        branches = ", ".join(b.name for b in e.branches.all()) if e.branches.exists() else "-"
        rows.append([i,e.first_name,e.last_name,e.phone,e.position.name if e.position else "-",e.role.name if e.role else "-",branches,e.created_at.strftime('%d.%m.%Y') if e.created_at else '-'])
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="xodimlar.csv"'
    return _write_csv(response, headers, rows)

@login_required(login_url="login")
def pending_export_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    students = Student.objects.filter(groups__isnull=True, status="kutilyotgan").select_related("desired_course").order_by("-created_at")
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Kutilyotganlar"
    headers = ["#","Ism","Familya","Telefon","Istagan kursi","Qo'shilgan sana"]
    hf = Font(bold=True,color="FFFFFF",size=11); hfl = PatternFill(start_color="2563EB",end_color="2563EB",fill_type="solid")
    tb = Border(left=Side(style='thin',color='D1D5DB'),right=Side(style='thin',color='D1D5DB'),top=Side(style='thin',color='D1D5DB'),bottom=Side(style='thin',color='D1D5DB'))
    for c,h in enumerate(headers,1): cell = ws.cell(row=1,column=c,value=h); cell.font = hf; cell.fill = hfl; cell.alignment = Alignment(horizontal='center',vertical='center'); cell.border = tb
    for i,s in enumerate(students,1):
        row = [i,s.first_name,s.last_name,s.phone or "-",s.desired_course.name if s.desired_course else "-",s.created_at.strftime('%d.%m.%Y') if s.created_at else '-']
        for c,v in enumerate(row,1): cell = ws.cell(row=i+1,column=c,value=v); cell.border = tb; cell.alignment = Alignment(vertical='center')
    for col,w in [(1,5),(2,18),(3,18),(4,20),(5,22),(6,14)]: ws.column_dimensions[chr(64+col)].width = w
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="kutilyotganlar.xlsx"'; wb.save(response); return response

@login_required(login_url="login")
def pending_export_csv(request):
    students = Student.objects.filter(groups__isnull=True, status="kutilyotgan").select_related("desired_course").order_by("-created_at")
    headers = ["#","Ism","Familya","Telefon","Istagan kursi","Qo'shilgan sana"]
    rows = [[i,s.first_name,s.last_name,s.phone or "-",s.desired_course.name if s.desired_course else "-",s.created_at.strftime('%d.%m.%Y') if s.created_at else '-'] for i,s in enumerate(students,1)]
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="kutilyotganlar.csv"'
    return _write_csv(response, headers, rows)

@login_required(login_url="login")
def graduated_export_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    students = Student.objects.filter(graduated_groups__isnull=False).prefetch_related("graduated_groups","groups").distinct().order_by("-created_at")
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Bitirilganlar"
    headers = ["#","Ism","Familya","Telefon","Bitirgan guruhlari","Hozirgi guruhlari","Qo'shilgan sana"]
    hf = Font(bold=True,color="FFFFFF",size=11); hfl = PatternFill(start_color="2563EB",end_color="2563EB",fill_type="solid")
    tb = Border(left=Side(style='thin',color='D1D5DB'),right=Side(style='thin',color='D1D5DB'),top=Side(style='thin',color='D1D5DB'),bottom=Side(style='thin',color='D1D5DB'))
    for c,h in enumerate(headers,1): cell = ws.cell(row=1,column=c,value=h); cell.font = hf; cell.fill = hfl; cell.alignment = Alignment(horizontal='center',vertical='center'); cell.border = tb
    for i,s in enumerate(students,1):
        grad = ", ".join(g.name for g in s.graduated_groups.all()) or "-"
        curr = ", ".join(g.name for g in s.groups.all()) or "-"
        row = [i,s.first_name,s.last_name,s.phone or "-",grad,curr,s.created_at.strftime('%d.%m.%Y') if s.created_at else '-']
        for c,v in enumerate(row,1): cell = ws.cell(row=i+1,column=c,value=v); cell.border = tb; cell.alignment = Alignment(vertical='center')
    for col,w in [(1,5),(2,18),(3,18),(4,20),(5,35),(6,25),(7,14)]: ws.column_dimensions[chr(64+col)].width = w
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="bitirilganlar.xlsx"'; wb.save(response); return response

@login_required(login_url="login")
def graduated_export_csv(request):
    students = Student.objects.filter(graduated_groups__isnull=False).prefetch_related("graduated_groups","groups").distinct().order_by("-created_at")
    headers = ["#","Ism","Familya","Telefon","Bitirgan guruhlari","Hozirgi guruhlari","Qo'shilgan sana"]
    rows = []
    for i,s in enumerate(students,1):
        grad = ", ".join(g.name for g in s.graduated_groups.all()) or "-"
        curr = ", ".join(g.name for g in s.groups.all()) or "-"
        rows.append([i,s.first_name,s.last_name,s.phone or "-",grad,curr,s.created_at.strftime('%d.%m.%Y') if s.created_at else '-'])
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="bitirilganlar.csv"'
    return _write_csv(response, headers, rows)


@login_required(login_url="login")
def group_detail_export_excel(request, pk):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    group = get_object_or_404(Group, pk=pk)
    students = group.students.prefetch_related("groups").annotate(total_groups=Count("groups"))
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = f"Guruh {group.name}"
    headers = ["#","Ism","Familya","Telefon","Guruhlari","Guruhlar soni","Qo'shilgan sana"]
    hf = Font(bold=True,color="FFFFFF",size=11); hfl = PatternFill(start_color="2563EB",end_color="2563EB",fill_type="solid")
    tb = Border(left=Side(style='thin',color='D1D5DB'),right=Side(style='thin',color='D1D5DB'),top=Side(style='thin',color='D1D5DB'),bottom=Side(style='thin',color='D1D5DB'))
    for c,h in enumerate(headers,1): cell = ws.cell(row=1,column=c,value=h); cell.font = hf; cell.fill = hfl; cell.alignment = Alignment(horizontal='center',vertical='center'); cell.border = tb
    for i,s in enumerate(students,1):
        groups_str = ", ".join(g.name for g in s.groups.all()) if s.groups.exists() else "-"
        row = [i,s.first_name,s.last_name,s.phone or "-",groups_str,s.total_groups,s.created_at.strftime('%d.%m.%Y') if s.created_at else '-']
        for c,v in enumerate(row,1): cell = ws.cell(row=i+1,column=c,value=v); cell.border = tb; cell.alignment = Alignment(vertical='center')
    for col,w in [(1,5),(2,18),(3,18),(4,20),(5,35),(6,12),(7,14)]: ws.column_dimensions[chr(64+col)].width = w
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="guruh_{group.name}_oquvchilar.xlsx"'; wb.save(response); return response

@login_required(login_url="login")
def group_detail_export_csv(request, pk):
    group = get_object_or_404(Group, pk=pk)
    students = group.students.prefetch_related("groups").annotate(total_groups=Count("groups"))
    headers = ["#","Ism","Familya","Telefon","Guruhlari","Guruhlar soni","Qo'shilgan sana"]
    rows = []
    for i,s in enumerate(students,1):
        groups_str = ", ".join(g.name for g in s.groups.all()) if s.groups.exists() else "-"
        rows.append([i,s.first_name,s.last_name,s.phone or "-",groups_str,s.total_groups,s.created_at.strftime('%d.%m.%Y') if s.created_at else '-'])
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="guruh_{group.name}_oquvchilar.csv"'
    return _write_csv(response, headers, rows)


@login_required(login_url="login")
def room_export_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    rooms = Room.objects.all().order_by("-created_at")
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Xonalar"
    headers = ["#","Xona nomi","Yaratilgan sana"]
    hf = Font(bold=True,color="FFFFFF",size=11); hfl = PatternFill(start_color="2563EB",end_color="2563EB",fill_type="solid")
    tb = Border(left=Side(style='thin',color='D1D5DB'),right=Side(style='thin',color='D1D5DB'),top=Side(style='thin',color='D1D5DB'),bottom=Side(style='thin',color='D1D5DB'))
    for c,h in enumerate(headers,1): cell = ws.cell(row=1,column=c,value=h); cell.font = hf; cell.fill = hfl; cell.alignment = Alignment(horizontal='center',vertical='center'); cell.border = tb
    for i,r in enumerate(rooms,1):
        row = [i,r.name,r.created_at.strftime('%d.%m.%Y') if r.created_at else '-']
        for c,v in enumerate(row,1): cell = ws.cell(row=i+1,column=c,value=v); cell.border = tb; cell.alignment = Alignment(vertical='center')
    for col,w in [(1,5),(2,25),(3,14)]: ws.column_dimensions[chr(64+col)].width = w
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="xonalar.xlsx"'; wb.save(response); return response

@login_required(login_url="login")
def room_export_csv(request):
    rooms = Room.objects.all().order_by("-created_at")
    headers = ["#","Xona nomi","Yaratilgan sana"]
    rows = [[i,r.name,r.created_at.strftime('%d.%m.%Y') if r.created_at else '-'] for i,r in enumerate(rooms,1)]
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="xonalar.csv"'
    return _write_csv(response, headers, rows)


@login_required(login_url="login")
def student_create(request):
    form = StudentCreateForm()
    if request.method == "POST":
        form = StudentCreateForm(request.POST)
        if form.is_valid():
            student = form.save()
            groups = form.cleaned_data.get("groups")
            if groups:
                student.groups.set(groups)
                first_group = groups[0]
                messages.success(request, f"O'quvchi {first_group.name} guruhiga qo'shildi")
                return redirect("group_detail", pk=first_group.pk)
            else:
                messages.success(request, "O'quvchi kutilyotganlar ro'yxatiga qo'shildi")
                return redirect("pending_students")
    else:
        initial_group = request.GET.get("group")
        if initial_group:
            form = StudentCreateForm(initial={"groups": [initial_group]})
    group_data = list(Group.objects.filter(status__in=["aktiv", "kutilyotgan"]).values("pk", "name", "course_id"))
    return render(request, "student/form.html", {"form": form, "title": "O'quvchi qo'shish", "group_data": group_data})


@login_required(login_url="login")
def student_update(request, pk):
    student = get_object_or_404(Student, pk=pk)
    form = StudentEditForm(instance=student)
    if request.method == "POST":
        form = StudentEditForm(request.POST, instance=student)
        if form.is_valid():
            form.save()
            messages.success(request, "O'quvchi muvaffaqiyatli yangilandi")
            return redirect("student_list")
    group_data = list(Group.objects.filter(status__in=["aktiv", "kutilyotgan"]).values("pk", "name", "course_id"))
    return render(request, "student/form.html", {"form": form, "title": "O'quvchini tahrirlash", "group_data": group_data})


@login_required(login_url="login")
def student_delete(request, pk):
    student = get_object_or_404(Student, pk=pk)
    if request.method == "POST":
        student.delete()
        messages.success(request, "O'quvchi muvaffaqiyatli o'chirildi")
        return redirect("student_list")
    return render(request, "student/delete.html", {"object": student, "title": "O'quvchini o'chirish"})


@login_required(login_url="login")
def pending_students(request):
    students = Student.objects.filter(groups__isnull=True, status="kutilyotgan").select_related("marketing_survey", "desired_course").order_by("desired_course__name", "-created_at")
    return render(request, "student/pending.html", {"students": students})


@login_required(login_url="login")
def group_detail(request, pk):
    group = get_object_or_404(Group.objects.annotate(
        total_students=Count("students")
    ).select_related("course", "room", "teacher").prefetch_related("lesson_times"), pk=pk)
    students = group.students.prefetch_related("groups").annotate(total_groups=Count("groups"))
    q = request.GET.get("q", "").strip()
    all_students = Student.objects.exclude(pk__in=students.values_list("pk", flat=True))
    if q:
        all_students = all_students.filter(
            Q(first_name__icontains=q) |
            Q(last_name__icontains=q) |
            Q(phone__endswith=q)
        )
    all_students = all_students.prefetch_related("groups").order_by("-created_at")
    pending_students = Student.objects.filter(
        groups__isnull=True, status="kutilyotgan"
    ).select_related("desired_course").prefetch_related("groups").order_by("-created_at")
    courses = Course.objects.all()
    removed_logs = StudentLog.objects.filter(group=group, action="removed").select_related("student").order_by("-created_at")[:50]
    frozen_students = group.students.filter(frozen_until__gte=date.today()).prefetch_related("groups")
    graduated_students = group.graduated_students.all().prefetch_related("groups")

    # Attendance data
    weekday_map = {0:"dushanba",1:"seshanba",2:"chorshanba",3:"payshanba",4:"juma",5:"shanba",6:"yakshanba"}
    today_uz = weekday_map[date.today().weekday()]
    today_lesson = group.lesson_times.filter(days__contains=today_uz).first()
    today_attendances = Attendance.objects.filter(group=group, date=date.today())
    attendance_map = {a.student_id: a.status for a in today_attendances}
    attendance_notes = {a.student_id: a.notes for a in today_attendances if a.notes}

    absence_reasons = AbsenceReason.objects.filter(is_active=True).order_by("order", "name")

    return render(request, "group/detail.html", {
        "group": group,
        "students": students,
        "all_students": all_students,
        "pending_students": pending_students,
        "removed_logs": removed_logs,
        "frozen_students": frozen_students,
        "graduated_students": graduated_students,
        "courses": courses,
        "q": q,
        "today_lesson": today_lesson,
        "attendance_map": attendance_map,
        "attendance_notes": attendance_notes,
        "absence_reasons": absence_reasons,
    })


def _add_student_to_group(student, group, reason="Guruh sahifasidan qo'shildi"):
    if group in student.groups.all():
        return False
    student.groups.add(group)
    student.frozen_until = None
    student.status = "kutilyotgan"
    student.save(update_fields=["frozen_until", "status"])
    StudentLog.objects.create(
        student=student, group=group, action="joined", reason=reason
    )
    return True


@login_required(login_url="login")
def add_student_to_group(request, group_pk, student_pk):
    group = get_object_or_404(Group, pk=group_pk)
    student = get_object_or_404(Student, pk=student_pk)
    if not _add_student_to_group(student, group):
        messages.warning(request, f"{student.first_name} {student.last_name} allaqachon {group.name} guruhiga qo'shilgan!")
        return redirect("group_detail", pk=group_pk)
    messages.success(request, f"{student.first_name} {student.last_name} {group.name} guruhiga qo'shildi")
    return redirect("group_detail", pk=group_pk)


@login_required(login_url="login")
def add_pending_to_group(request, group_pk, course_pk):
    group = get_object_or_404(Group, pk=group_pk)
    course = get_object_or_404(Course, pk=course_pk)
    students = Student.objects.filter(groups__isnull=True, status="kutilyotgan", desired_course=course)
    count = 0
    for student in students:
        if _add_student_to_group(student, group, f"{group.name} guruhiga kurs bo'yicha qo'shildi"):
            count += 1
    if count:
        messages.success(request, f"{count} ta o'quvchi {group.name} guruhiga qo'shildi")
    else:
        messages.info(request, "Qo'shiladigan o'quvchi topilmadi")
    return redirect("group_detail", pk=group_pk)


@login_required(login_url="login")
def remove_student_from_group(request, group_pk, student_pk):
    group = get_object_or_404(Group, pk=group_pk)
    student = get_object_or_404(Student, pk=student_pk)
    if request.method == "POST":
        reason = request.POST.get("reason", "").strip()
        if not reason:
            messages.error(request, "Chiqarish sababini yozing!")
            return redirect("group_detail", pk=group_pk)
        student.groups.remove(group)
        student.frozen_until = None
        student.status = "chiqarilgan"
        student.save(update_fields=["frozen_until", "status"])
        group_count = student.groups.count()
        if group_count == 0:
            student.status = "chiqarilgan"
            student.save(update_fields=["status"])
        StudentLog.objects.create(
            student=student, group=group, action="removed", reason=reason
        )
        messages.success(request, f"{student.first_name} {student.last_name} guruhdan chiqarildi")
        return redirect("group_detail", pk=group_pk)
    return redirect("group_detail", pk=group_pk)


@login_required(login_url="login")
def graduate_student(request, group_pk, student_pk):
    group = get_object_or_404(Group, pk=group_pk)
    student = get_object_or_404(Student, pk=student_pk)
    if group not in student.groups.all():
        messages.warning(request, f"{student.first_name} {student.last_name} bu guruhda emas!")
        return redirect("group_detail", pk=group_pk)
    if group in student.graduated_groups.all():
        messages.warning(request, f"{student.first_name} {student.last_name} allaqachon {group.name} dan bitirilgan!")
        return redirect("group_detail", pk=group_pk)
    student.groups.remove(group)
    student.graduated_groups.add(group)
    StudentLog.objects.create(
        student=student, group=group, action="graduated",
        reason=f"{group.name} guruhini bitirdi"
    )
    messages.success(request, f"{student.first_name} {student.last_name} {group.name} guruhini bitirdi!")
    return redirect("group_detail", pk=group_pk)


@login_required(login_url="login")
def transfer_student(request, group_pk, student_pk):
    group = get_object_or_404(Group, pk=group_pk)
    student = get_object_or_404(Student, pk=student_pk)
    if group not in student.groups.all():
        messages.warning(request, f"{student.first_name} {student.last_name} bu guruhda emas!")
        return redirect("group_detail", pk=group_pk)

    if request.method == "POST":
        reason = request.POST.get("reason", "").strip()
        new_group_id = request.POST.get("new_group")
        if new_group_id == "pending":
            old_name = group.name
            student.groups.remove(group)
            student.frozen_until = None
            student.status = "kutilyotgan"
            student.save(update_fields=["frozen_until", "status"])
            StudentLog.objects.create(
                student=student, group=group, action="transferred",
                reason=reason or f"{old_name} → Kutilyotganlar"
            )
            messages.success(request, f"{student.first_name} {student.last_name} {old_name} dan kutilyotganlarga o'tkazildi!")
            return redirect("pending_students")
        if not new_group_id:
            messages.error(request, "Yangi guruhni tanlang!")
            return redirect("transfer_student", group_pk=group_pk, student_pk=student_pk)
        new_group = get_object_or_404(Group, pk=new_group_id)
        trans_info = f"{group.name} → {new_group.name}"
        full_reason = f"{reason} | {trans_info}" if reason else trans_info
        student.groups.remove(group)
        student.groups.add(new_group)
        student.frozen_until = None
        student.status = "kutilyotgan"
        student.save(update_fields=["frozen_until", "status"])
        StudentLog.objects.create(
            student=student, group=group, action="transferred",
            reason=full_reason
        )
        messages.success(request, f"{student.first_name} {student.last_name} {group.name} dan {new_group.name} ga o'tkazildi!")
        return redirect("group_detail", pk=new_group.pk)

    course_id = request.GET.get("course_id")
    teacher_id = request.GET.get("teacher_id")

    courses = Course.objects.all().order_by("name")
    teachers = Employee.objects.none()
    groups = Group.objects.none()

    if course_id:
        teachers = Employee.objects.filter(
            role__name="O'qituvchi",
            teacher_groups__course_id=course_id
        ).exclude(teacher_groups__isnull=True).distinct().order_by("first_name")

    if teacher_id and course_id:
        groups = Group.objects.filter(
            status__in=["aktiv", "kutilyotgan"],
            course_id=course_id,
            teacher_id=teacher_id,
        ).exclude(pk=group.pk).select_related(
            "course", "room", "teacher"
        ).prefetch_related("lesson_times").order_by("name")

    return render(request, "student/transfer.html", {
        "student": student,
        "group": group,
        "courses": courses,
        "teachers": teachers,
        "groups": groups,
        "selected_course_id": course_id,
        "selected_teacher_id": teacher_id,
    })


@login_required(login_url="login")
def transfer_all_students(request, pk):
    group = get_object_or_404(Group, pk=pk)
    students = group.students.all()
    if request.method == "POST":
        reason = request.POST.get("reason", "").strip()
        new_group_id = request.POST.get("new_group")
        if not new_group_id or new_group_id == "pending":
            messages.error(request, "Yangi guruhni tanlang!")
            return redirect("transfer_all_students", pk=pk)
        new_group = get_object_or_404(Group, pk=new_group_id)
        trans_info = f"{group.name} → {new_group.name}"
        full_reason = f"{reason} | {trans_info}" if reason else trans_info
        count = 0
        for student in students:
            student.groups.remove(group)
            student.groups.add(new_group)
            student.frozen_until = None
            student.status = "kutilyotgan"
            student.save(update_fields=["frozen_until", "status"])
            StudentLog.objects.create(
                student=student, group=group, action="transferred",
                reason=full_reason
            )
            count += 1
        messages.success(request, f"{count} ta o'quvchi {group.name} dan {new_group.name} ga o'tkazildi!")
        return redirect("group_detail", pk=new_group.pk)
    course_id = request.GET.get("course_id")
    teacher_id = request.GET.get("teacher_id")
    courses = Course.objects.all().order_by("name")
    teachers = Employee.objects.none()
    groups = Group.objects.none()
    if course_id:
        teachers = Employee.objects.filter(
            role__name="O'qituvchi", teacher_groups__course_id=course_id
        ).exclude(teacher_groups__isnull=True).distinct().order_by("first_name")
    if teacher_id and course_id:
        groups = Group.objects.filter(
            status__in=["aktiv", "kutilyotgan"],
            course_id=course_id, teacher_id=teacher_id,
        ).exclude(pk=group.pk).select_related("course", "room", "teacher"
        ).prefetch_related("lesson_times").annotate(
            student_count=Count("students")
        ).order_by("name")
    return render(request, "group/transfer_all.html", {
        "group": group, "students": students, "courses": courses,
        "teachers": teachers, "groups": groups,
        "selected_course_id": course_id, "selected_teacher_id": teacher_id,
    })


@login_required(login_url="login")
def graduated_students(request):
    students = Student.objects.filter(graduated_groups__isnull=False).prefetch_related(
        "graduated_groups", "groups"
    ).distinct().order_by("-created_at")
    return render(request, "student/graduated.html", {"students": students})


@login_required(login_url="login")
def student_profile(request, pk):
    student = get_object_or_404(Student.objects.prefetch_related("groups", "graduated_groups"), pk=pk)
    logs = list(student.logs.select_related("group").all())
    import re
    for log in logs:
        log.target_name = None
        if log.action == "transferred" and log.reason:
            m = re.search(r"→\s*(.+?)$", log.reason)
            if m:
                log.target_name = m.group(1).strip()
            if not log.target_name:
                m = re.search(r"\bdan\b\s+(.+?)\s+ga\s+o'tkazildi", log.reason)
                if m:
                    log.target_name = m.group(1).strip()
            if not log.target_name and re.search(r"\bkutilyotgan", log.reason, re.I):
                log.target_name = "Kutilyotganlar"
    groups = Group.objects.filter(status="aktiv").order_by("name")

    # Attendance history for this student
    attendance_history = Attendance.objects.filter(
        student=student
    ).select_related("group", "teacher").order_by("-date")[:30]

    return render(request, "student/profile.html", {
        "student": student, "logs": logs, "groups": groups,
        "attendance_history": attendance_history,
    })


@login_required(login_url="login")
def student_freeze(request, pk):
    student = get_object_or_404(Student, pk=pk)
    if request.method == "POST":
        form = FreezeForm(request.POST)
        if form.is_valid():
            days = form.cleaned_data["days"]
            reason = form.cleaned_data["reason"]
            frozen_until = date.today() + timedelta(days=days)
            student.frozen_until = frozen_until
            student.save(update_fields=["frozen_until"])
            StudentLog.objects.create(
                student=student, group=student.groups.first(), action="frozen",
                reason=f"{days} kunga muzlatildi. {reason}" if reason else f"{days} kunga muzlatildi"
            )

            # Auto-create absent attendance for future lesson dates
            freez_note = "Muzlatilgan"
            if reason:
                freez_note += f" - {reason}"
            weekday_map = {0:"dushanba",1:"seshanba",2:"chorshanba",3:"payshanba",4:"juma",5:"shanba",6:"yakshanba"}
            today = date.today()
            for group in student.groups.all():
                lesson_day_nums = set()
                for lt in group.lesson_times.all():
                    for d_name in lt.days.split(","):
                        d_name = d_name.strip()
                        for num, uz_name in weekday_map.items():
                            if uz_name == d_name:
                                lesson_day_nums.add(num)
                d = today
                while d <= frozen_until:
                    if d.weekday() in lesson_day_nums:
                        Attendance.objects.update_or_create(
                            group=group,
                            student=student,
                            date=d,
                            defaults={"status": "absent", "notes": freez_note}
                        )
                    d += timedelta(days=1)

            messages.success(request, f"{student.first_name} {student.last_name} {days} kunga muzlatildi")
            return redirect("student_profile", pk=student.pk)
    else:
        form = FreezeForm()
    return render(request, "student/freeze.html", {"form": form, "student": student})


@login_required(login_url="login")
def student_unfreeze(request, pk):
    student = get_object_or_404(Student, pk=pk)
    # Remove auto-created absent records from freezing
    Attendance.objects.filter(
        student=student,
        status="absent",
        notes__startswith="Muzlatilgan"
    ).delete()
    student.frozen_until = None
    student.save(update_fields=["frozen_until"])
    StudentLog.objects.create(
        student=student, group=student.groups.first(), action="unfrozen",
        reason="Muzlatish bekor qilindi"
    )
    messages.success(request, f"{student.first_name} {student.last_name} muzlatish bekor qilindi")
    return redirect("student_profile", pk=student.pk)


@login_required(login_url="login")
def student_remove_from_group(request, pk):
    student = get_object_or_404(Student, pk=pk)
    if request.method == "POST":
        form = RemoveFromGroupForm(request.POST)
        if form.is_valid():
            reason = form.cleaned_data["reason"]
            group = student.groups.first()
            student.groups.clear()
            student.frozen_until = None
            student.status = "chiqarilgan"
            student.save(update_fields=["frozen_until", "status"])
            StudentLog.objects.create(
                student=student, group=group, action="removed", reason=reason
            )
            messages.success(request, f"{student.first_name} {student.last_name} guruhdan chiqarildi")
            if group:
                return redirect("group_detail", pk=group.pk)
            return redirect("student_profile", pk=student.pk)
    else:
        form = RemoveFromGroupForm()
    return render(request, "student/remove.html", {"form": form, "student": student})


@login_required(login_url="login")
def student_add_to_group(request, pk):
    student = get_object_or_404(Student, pk=pk)
    if request.method == "POST":
        form = AddToGroupForm(request.POST)
        if form.is_valid():
            group = form.cleaned_data["group"]
            if group in student.groups.all():
                messages.warning(request, f"{student.first_name} {student.last_name} allaqachon {group.name} guruhiga qo'shilgan!")
                return redirect("student_profile", pk=student.pk)
            reason = form.cleaned_data["reason"]
            student.groups.add(group)
            student.frozen_until = None
            student.status = "kutilyotgan"
            student.save(update_fields=["frozen_until", "status"])
            StudentLog.objects.create(
                student=student, group=group, action="joined", reason=reason or "Profil sahifasidan qo'shildi"
            )
            messages.success(request, f"{student.first_name} {student.last_name} {group.name} guruhiga qo'shildi")
            return redirect("group_detail", pk=group.pk)
    else:
        form = AddToGroupForm()
    return render(request, "student/add_to_group.html", {"form": form, "student": student})


@login_required(login_url="login")
def group_freeze(request, pk):
    group = get_object_or_404(Group, pk=pk)
    if request.method == "POST":
        form = FreezeForm(request.POST)
        if form.is_valid():
            days = form.cleaned_data["days"]
            reason = form.cleaned_data["reason"]
            frozen_until = date.today() + timedelta(days=days)
            students = group.students.filter(
                Q(frozen_until__isnull=True) | Q(frozen_until__lt=date.today())
            )
            for student in students:
                student.frozen_until = frozen_until
                student.save(update_fields=["frozen_until"])
                StudentLog.objects.create(
                    student=student, group=group, action="frozen",
                    reason=f"Guruh bilan {days} kunga muzlatildi. {reason}" if reason else f"Guruh bilan {days} kunga muzlatildi"
                )
            messages.success(request, f"Guruh {days} kunga muzlatildi ({students.count()} ta o'quvchi)")
            return redirect("group_detail", pk=group.pk)
    else:
        form = FreezeForm()
    return render(request, "group/freeze.html", {"form": form, "group": group})


@login_required(login_url="login")
def group_archive(request, pk):
    group = get_object_or_404(Group, pk=pk)
    if request.method == "POST":
        students = group.students.all()
        for student in students:
            student.groups.remove(group)
            student.graduated_groups.add(group)
            student.frozen_until = None
            student.status = "bitirilgan"
            student.save(update_fields=["frozen_until", "status"])
            StudentLog.objects.create(
                student=student, group=group, action="graduated",
                reason=f"{group.name} guruhi arxivlandi"
            )
        group.status = "arxivlangan"
        group.save()
        messages.success(request, f"Guruh arxivlandi va {students.count()} ta o'quvchi bitirildi")
        return redirect("group_list")
    return render(request, "group/archive.html", {"group": group})


@login_required(login_url="login")
def group_settings(request, pk):
    group = get_object_or_404(Group, pk=pk)
    group_form = GroupForm(instance=group)
    lesson_form = LessonTimeForm(group=group)
    lesson_times = group.lesson_times.all()

    if request.method == "POST":
        if "update_group" in request.POST:
            group_form = GroupForm(request.POST, instance=group)
            if group_form.is_valid():
                group_form.save()
                messages.success(request, "Guruh sozlamalari saqlandi")
                return redirect("group_settings", pk=group.pk)
        elif "add_lesson" in request.POST:
            lesson_form = LessonTimeForm(request.POST, group=group)
            if lesson_form.is_valid():
                lesson = lesson_form.save(commit=False)
                lesson.group = group
                lesson.save()
                messages.success(request, "Dars vaqti qo'shildi")
                return redirect("group_settings", pk=group.pk)
        elif "delete_lesson" in request.POST:
            lesson_id = request.POST.get("lesson_id")
            if lesson_id:
                LessonTime.objects.filter(pk=lesson_id, group=group).delete()
                messages.success(request, "Dars vaqti o'chirildi")
                return redirect("group_settings", pk=group.pk)

    return render(request, "group/settings.html", {
        "group": group,
        "group_form": group_form,
        "lesson_form": lesson_form,
        "lesson_times": lesson_times,
    })


@login_required(login_url="login")
def employee_list(request):
    employees = Employee.objects.select_related("position", "role").all().order_by("-created_at")
    return render(request, "employee/list.html", {"employees": employees})


@login_required(login_url="login")
def employee_create(request):
    form = EmployeeForm()
    if request.method == "POST":
        form = EmployeeForm(request.POST, request.FILES)
        if form.is_valid():
            employee = form.save(commit=False)
            password = form.cleaned_data.get("password")
            if password:
                user = User.objects.create_user(
                    username=employee.phone,
                    password=password,
                    first_name=employee.first_name,
                    last_name=employee.last_name,
                )
                employee.user = user
            employee.save()
            form.save_m2m()
            messages.success(request, "Xodim muvaffaqiyatli qo'shildi")
            return redirect("employee_list")
    return render(request, "employee/create.html", {"form": form})


@login_required(login_url="login")
def branch_list(request):
    branches = Branch.objects.all().order_by("-created_at")
    return render(request, "branch/list.html", {"branches": branches})


@login_required(login_url="login")
def branch_create(request):
    form = BranchForm()
    if request.method == "POST":
        form = BranchForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Filial muvaffaqiyatli qo'shildi")
            return redirect("branch_list")
    return render(request, "branch/form.html", {"form": form, "title": "Filial qo'shish"})


@login_required(login_url="login")
def branch_update(request, pk):
    branch = get_object_or_404(Branch, pk=pk)
    form = BranchForm(instance=branch)
    if request.method == "POST":
        form = BranchForm(request.POST, instance=branch)
        if form.is_valid():
            form.save()
            messages.success(request, "Filial muvaffaqiyatli yangilandi")
            return redirect("branch_list")
    return render(request, "branch/form.html", {"form": form, "title": "Filialni tahrirlash"})


@login_required(login_url="login")
def branch_delete(request, pk):
    branch = get_object_or_404(Branch, pk=pk)
    if request.method == "POST":
        branch.delete()
        messages.success(request, "Filial muvaffaqiyatli o'chirildi")
        return redirect("branch_list")
    return render(request, "branch/delete.html", {"object": branch, "title": "Filialni o'chirish"})


@login_required(login_url="login")
def room_list(request):
    rooms = Room.objects.all().order_by("-created_at")
    return render(request, "room/list.html", {"rooms": rooms})


@login_required(login_url="login")
def room_create(request):
    form = RoomForm()
    if request.method == "POST":
        form = RoomForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Xona muvaffaqiyatli qo'shildi")
            return redirect("room_list")
    return render(request, "room/form.html", {"form": form, "title": "Xona qo'shish"})


@login_required(login_url="login")
def room_update(request, pk):
    room = get_object_or_404(Room, pk=pk)
    form = RoomForm(instance=room)
    if request.method == "POST":
        form = RoomForm(request.POST, instance=room)
        if form.is_valid():
            form.save()
            messages.success(request, "Xona muvaffaqiyatli yangilandi")
            return redirect("room_list")
    return render(request, "room/form.html", {"form": form, "title": "Xonani tahrirlash"})


@login_required(login_url="login")
def room_delete(request, pk):
    room = get_object_or_404(Room, pk=pk)
    if request.method == "POST":
        room.delete()
        messages.success(request, "Xona muvaffaqiyatli o'chirildi")
        return redirect("room_list")
    return render(request, "room/delete.html", {"object": room, "title": "Xonani o'chirish"})


@login_required(login_url="login")
def position_list(request):
    positions = Position.objects.all().order_by("-created_at")
    return render(request, "position/list.html", {"positions": positions})


@login_required(login_url="login")
def position_create(request):
    form = PositionForm()
    if request.method == "POST":
        form = PositionForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Vazifa muvaffaqiyatli qo'shildi")
            return redirect("position_list")
    return render(request, "position/form.html", {"form": form, "title": "Vazifa qo'shish"})


@login_required(login_url="login")
def position_update(request, pk):
    position = get_object_or_404(Position, pk=pk)
    form = PositionForm(instance=position)
    if request.method == "POST":
        form = PositionForm(request.POST, instance=position)
        if form.is_valid():
            form.save()
            messages.success(request, "Vazifa muvaffaqiyatli yangilandi")
            return redirect("position_list")
    return render(request, "position/form.html", {"form": form, "title": "Vazifani tahrirlash"})


@login_required(login_url="login")
def position_delete(request, pk):
    position = get_object_or_404(Position, pk=pk)
    if request.method == "POST":
        position.delete()
        messages.success(request, "Vazifa muvaffaqiyatli o'chirildi")
        return redirect("position_list")
    return render(request, "position/delete.html", {"object": position, "title": "Vazifani o'chirish"})


@login_required(login_url="login")
def statistics(request):
    total_courses = Course.objects.count()
    total_groups = Group.objects.count()
    total_students = Student.objects.count()
    total_surveys = MarketingSurvey.objects.count()
    pending_students = Student.objects.filter(groups__isnull=True, status="kutilyotgan").count()
    active_groups = Group.objects.filter(status="aktiv").count()
    pending_groups_count = Group.objects.filter(status="kutilyotgan").count()
    online_groups = Group.objects.filter(education_type="onlayn").count()
    offline_groups = Group.objects.filter(education_type="oflayn").count()
    toq_groups = Group.objects.filter(day_type="toq").count()
    juft_groups = Group.objects.filter(day_type="juft").count()
    har_kun_groups = Group.objects.filter(day_type="har_kun").count()

    group_students = Group.objects.annotate(count=Count("students")).values("name", "count")
    survey_stats = MarketingSurvey.objects.annotate(
        count=Count("students")
    ).order_by("-count")
    total_survey_students = sum(s.count for s in survey_stats)

    recent_students = Student.objects.prefetch_related("groups", "marketing_survey").order_by("-created_at")[:10]

    return render(request, "statistics.html", {
        "total_courses": total_courses,
        "total_groups": total_groups,
        "total_students": total_students,
        "total_surveys": total_surveys,
        "pending_students": pending_students,
        "active_groups": active_groups,
        "pending_groups_count": pending_groups_count,
        "online_groups": online_groups,
        "offline_groups": offline_groups,
        "toq_groups": toq_groups,
        "juft_groups": juft_groups,
        "har_kun_groups": har_kun_groups,
        "group_students": group_students,
        "survey_stats": survey_stats,
        "total_survey_students": total_survey_students,
        "recent_students": recent_students,
    })


@login_required(login_url="login")
def survey_list(request):
    surveys = MarketingSurvey.objects.annotate(
        student_count=Count("students")
    ).order_by("-created_at")
    return render(request, "survey/list.html", {"surveys": surveys})


@login_required(login_url="login")
def survey_create(request):
    form = MarketingSurveyForm()
    if request.method == "POST":
        form = MarketingSurveyForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "So'rovnoma muvaffaqiyatli qo'shildi")
            return redirect("survey_list")
    return render(request, "survey/form.html", {"form": form, "title": "So'rovnoma qo'shish"})


@login_required(login_url="login")
def survey_update(request, pk):
    survey = get_object_or_404(MarketingSurvey, pk=pk)
    form = MarketingSurveyForm(instance=survey)
    if request.method == "POST":
        form = MarketingSurveyForm(request.POST, instance=survey)
        if form.is_valid():
            form.save()
            messages.success(request, "So'rovnoma muvaffaqiyatli yangilandi")
            return redirect("survey_list")
    return render(request, "survey/form.html", {"form": form, "title": "So'rovnomani tahrirlash"})


@login_required(login_url="login")
def survey_delete(request, pk):
    survey = get_object_or_404(MarketingSurvey, pk=pk)
    if request.method == "POST":
        survey.delete()
        messages.success(request, "So'rovnoma muvaffaqiyatli o'chirildi")
        return redirect("survey_list")
    return render(request, "survey/delete.html", {"object": survey, "title": "So'rovnomani o'chirish"})


@login_required(login_url="login")
def dismiss_removed_log(request, pk):
    if request.method == "POST":
        log = get_object_or_404(StudentLog, pk=pk)
        log.delete()
        return JsonResponse({"ok": True})
    return JsonResponse({"ok": False}, status=405)


@login_required(login_url="login")
def absence_reason_list(request):
    reasons = AbsenceReason.objects.all().order_by("order", "name")
    return render(request, "absence_reason/list.html", {"reasons": reasons})


@login_required(login_url="login")
def absence_reason_create(request):
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        reason_type = request.POST.get("reason_type", "both")
        is_active = request.POST.get("is_active") == "on"
        order = request.POST.get("order", 0)
        if name:
            AbsenceReason.objects.create(
                name=name,
                reason_type=reason_type,
                is_active=is_active,
                order=int(order) if order else 0,
            )
            messages.success(request, "Davomat sababi qo'shildi")
        else:
            messages.error(request, "Sabab nomini yozing!")
        return redirect("absence_reason_list")
    return redirect("absence_reason_list")


@login_required(login_url="login")
def absence_reason_update(request, pk):
    reason = get_object_or_404(AbsenceReason, pk=pk)
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        reason_type = request.POST.get("reason_type", "both")
        is_active = request.POST.get("is_active") == "on"
        order = request.POST.get("order", 0)
        if name:
            reason.name = name
            reason.reason_type = reason_type
            reason.is_active = is_active
            reason.order = int(order) if order else 0
            reason.save()
            messages.success(request, "Davomat sababi yangilandi")
        else:
            messages.error(request, "Sabab nomini yozing!")
        return redirect("absence_reason_list")
    return redirect("absence_reason_list")


@login_required(login_url="login")
def absence_reason_delete(request, pk):
    reason = get_object_or_404(AbsenceReason, pk=pk)
    if request.method == "POST":
        reason.delete()
        messages.success(request, "Davomat sababi o'chirildi")
        return redirect("absence_reason_list")
    return render(request, "absence_reason/delete.html", {"reason": reason})


# ---- Student Web Interface ----


